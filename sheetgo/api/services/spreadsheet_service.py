from openpyxl import load_workbook


class SpreadsheetServiceException(Exception):
    pass


class SpreadsheetService:
    """
        The idea here is to create a service that could be used with injector.
        This has only one method and no constructor because it only extract the list of tabs of the Excel xlsx file.
        In a real world application this service probably would have a bunch of other helpful methods.
    """
    @staticmethod
    def ordered_sheetnames(spreadsheet):
        try:
            wb = load_workbook(filename=spreadsheet)
        except Exception:
            raise SpreadsheetServiceException('could not load workbook')

        return sorted(wb.sheetnames)
